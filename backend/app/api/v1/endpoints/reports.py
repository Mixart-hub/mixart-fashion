from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date, timedelta
from typing import Optional
from app.db.database import get_db
from app.models.models import Order, OrderItem, Product, Category, User, Branch
import io

router = APIRouter()

def _date_range(period: str):
    today = date.today()
    if period == "today":
        return datetime.combine(today, datetime.min.time())
    elif period == "week":
        return datetime.combine(today - timedelta(days=7), datetime.min.time())
    else:
        return datetime.combine(today.replace(day=1), datetime.min.time())


@router.get("/summary")
def summary(
    period: str = "today",
    branch_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    start = _date_range(period)
    q = db.query(Order).filter(Order.created_at >= start)
    if branch_id:
        q = q.filter(Order.branch_id == branch_id)
    orders = q.all()

    revenue = sum(o.final_amount for o in orders if o.payment_status == "paid")
    total_revenue = sum(o.final_amount for o in orders)
    by_status = {}
    for o in orders:
        key = o.status.value if hasattr(o.status, "value") else str(o.status)
        by_status[key] = by_status.get(key, 0) + 1

    top_q = db.query(
        Product.name_uz, func.sum(OrderItem.quantity).label("qty")
    ).join(OrderItem, OrderItem.product_id == Product.id
    ).join(Order, Order.id == OrderItem.order_id
    ).filter(Order.created_at >= start)
    if branch_id:
        top_q = top_q.filter(Order.branch_id == branch_id)
    top_products = top_q.group_by(Product.name_uz
    ).order_by(func.sum(OrderItem.quantity).desc()
    ).limit(5).all()

    branch_name = None
    if branch_id:
        b = db.query(Branch).filter(Branch.id == branch_id).first()
        branch_name = b.name if b else None

    return {
        "period": period,
        "branch_id": branch_id,
        "branch_name": branch_name,
        "total_orders": len(orders),
        "total_revenue": round(total_revenue, 2),
        "paid_revenue": round(revenue, 2),
        "by_status": by_status,
        "top_products": [{"name": p.name_uz, "qty": p.qty} for p in top_products]
    }

@router.get("/excel")
def export_excel(period: str = "month", db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {"error": "openpyxl not installed"}

    today = date.today()
    start = datetime.combine(today.replace(day=1), datetime.min.time())
    orders = db.query(Order).filter(Order.created_at >= start).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="8B3A62")
    headers = ["#", "Sana", "Mijoz ID", "Summa", "Chegirma", "Jami", "To'lov", "Holat"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, o in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=o.id)
        ws.cell(row=row, column=2, value=str(o.created_at)[:10])
        ws.cell(row=row, column=3, value=o.customer_id)
        ws.cell(row=row, column=4, value=o.total_amount)
        ws.cell(row=row, column=5, value=o.discount_amount)
        ws.cell(row=row, column=6, value=o.final_amount)
        ws.cell(row=row, column=7, value=o.payment_status)
        ws.cell(row=row, column=8, value=o.status)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"mixart_report_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/pdf")
def export_pdf(period: str = "today", db: Session = Depends(get_db)):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
    except ImportError:
        return {"error": "reportlab not installed"}

    today = date.today()
    start = datetime.combine(today, datetime.min.time())
    orders = db.query(Order).filter(Order.created_at >= start).all()
    revenue = sum(o.final_amount for o in orders if o.payment_status == "paid")

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    c.setFillColorRGB(0.545, 0.227, 0.384)
    c.rect(0, h - 80, w, 80, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(40, h - 50, "MIXART Fashion")
    c.setFont("Helvetica", 12)
    c.drawString(40, h - 68, f"Kunlik hisobot — {today}")

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, h - 120, f"Jami buyurtmalar: {len(orders)}")
    c.drawString(40, h - 145, f"Jami daromad: ${revenue:.2f}")

    y = h - 200
    c.setFont("Helvetica-Bold", 11)
    c.drawString(40, y, "#"); c.drawString(80, y, "Sana"); c.drawString(180, y, "Summa"); c.drawString(280, y, "Holat")
    y -= 20
    c.setFont("Helvetica", 10)
    for o in orders[:30]:
        c.drawString(40, y, str(o.id))
        c.drawString(80, y, str(o.created_at)[:10])
        c.drawString(180, y, f"${o.final_amount:.2f}")
        c.drawString(280, y, str(o.status))
        y -= 18
        if y < 60:
            c.showPage()
            y = h - 60

    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=mixart_{today}.pdf"}
    )
